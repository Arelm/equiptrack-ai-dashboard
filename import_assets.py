"""EquipTrack asset importer.

Reads plot spreadsheets and creates Asset records, one per physical unit.

Layouts supported:

  RESIDENTIAL   Flat | Room | Make | Capacity   (165, 225, Cecilia Court, 266, 217A)
                Room | Make | Capacity          (262, 217B — no flat column)
  OFFICE        S/N | Block | Floor/Zone | Office/Area | Brand | Unit Type | HP | Qty
                                                (OML 13 A, OML 13 B)

The residential parser finds the header row and each column by name, so
leading blank columns and title rows in different places do not matter.

Repeated rows mean separate physical units in the same room — seven Halls
at 217B are seven machines. Where a generated name would repeat, the units
are numbered #1, #2, ... in sheet order.

DRY_RUN is True by default. Nothing is written until you set it to False.
Re-running is safe: an asset with the same name at the same location is
skipped, not duplicated.
"""

import os
import re
import sys
import uuid
from datetime import datetime, timezone

import psycopg2
from openpyxl import load_workbook

# =========================================================================
# CONFIG — one entry per sheet.
#
#   file / sheet   where to read
#   layout         "residential" or "office"
#   location       Location name in EquipTrack (created if missing)
#   prefix         what starts each asset name
#   area / client / supervisor / phone
#                  used when creating a Location, or to fill blanks on an
#                  existing one. Set FORCE_LOCATION_UPDATE below to
#                  overwrite values that are already there.
# =========================================================================

BIG = "217A_WhiteHouse___Cecilia_Court___262___266___217B.xlsx"

SOURCES = [
    # --- already imported; left here so re-runs stay consistent ---
    {"file": "165.xlsx", "sheet": "Location 165", "layout": "residential",
     "location": "165", "prefix": "165",
     "area": "V.I.", "client": "Sterling Oil", "supervisor": None, "phone": None},

    {"file": "flat_225.xlsx", "sheet": "Sheet2", "layout": "residential",
     "location": "225", "prefix": "225",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "Saadu Kaila", "phone": "07042246600"},

    {"file": "OML13_A___B.xlsx", "sheet": "OML 13 A", "layout": "office",
     "location": "OML 13 A", "prefix": "OML13A",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "Patrick", "phone": "08160746387"},

    {"file": "OML13_A___B.xlsx", "sheet": "OML 13 B", "layout": "office",
     "location": "OML 13 B", "prefix": "OML13B",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "Effiong", "phone": "09058693215"},

    # --- new ---
    {"file": BIG, "sheet": "217A", "layout": "residential",
     "location": "217A", "prefix": "217A",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "Sunny Bissong", "phone": "09065996593"},

    {"file": BIG, "sheet": "cecilia_Court", "layout": "residential",
     "location": "Cecilia Court", "prefix": "Cecilia",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "Joseph", "phone": "07046041957"},

    {"file": BIG, "sheet": "262", "layout": "residential",
     "location": "262", "prefix": "262",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": None, "phone": None},

    {"file": BIG, "sheet": "266", "layout": "residential",
     "location": "266", "prefix": "266",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "DamiLola", "phone": "08068870394"},

    {"file": BIG, "sheet": "217B", "layout": "residential",
     "location": "217B", "prefix": "217B",
     "area": "V.I.", "client": "Sterling Oil",
     "supervisor": "Martins", "phone": "08130002846"},
]

DRY_RUN = False                 # set False to write
FORCE_LOCATION_UPDATE = False  # True overwrites supervisor/area/client already set

# Residential sheets carry no unit-type column, so type is read from capacity:
#   under 3HP        Split AC
#   3HP to 10HP      Standing Unit
#   above 10HP       Package Unit   (rooftop / package plant)
STANDING_FROM_HP = 3.0
PACKAGE_ABOVE_HP = 10.0

SPELLING = {
    "kittchen": "Kitchen", "kitchen": "Kitchen",
    "dinning": "Dining", "dinniong": "Dining", "dining": "Dining",
    "palour": "Parlour", "parlour": "Parlour",
}

ZONE_SHORT = {
    "ground floor": "GF",
    "ground floor left": "GF LEFT",
    "ground floor right": "GF RIGHT",
    "1st floor left": "1F LEFT",
    "1st floor right": "1F RIGHT",
    "2nd floor left": "2F LEFT",
    "2nd floor right": "2F RIGHT",
}

# Header words used to locate columns.
H_FLAT = {"flat", "block", "zone"}
H_ROOM = {"room", "office / area", "office/area"}
H_MAKE = {"make", "ac make", "brand"}
H_CAP = {"capacity", "capacity (hp)", "hp"}


def clean(v):
    """Cell to a single-line trimmed string. Some cells contain newlines."""
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def fix_room(name):
    return SPELLING.get(name.strip().lower(), name.strip())


def short_zone(name):
    return ZONE_SHORT.get(name.strip().lower(), name.strip())


def fix_flat(raw):
    """'1' -> 'F1'. 'first Floor' -> 'First Floor'. 'BQ', 'A2' unchanged."""
    t = raw.strip()
    if t.replace(".", "").isdigit():
        return f"F{int(float(t))}"
    return " ".join(w[:1].upper() + w[1:] if w.islower() else w for w in t.split())


def hp_text(raw):
    t = raw.upper().replace("HP", "").strip()
    try:
        return f"{float(t):.1f}HP"
    except ValueError:
        return raw.strip()


def hp_value(raw):
    t = raw.upper().replace("HP", "").strip()
    try:
        return float(t)
    except ValueError:
        return 0.0


def find_header(rows):
    """Locate the header row and the column index of each field.

    Flat is optional — 262 and 217B list rooms with no flat column.
    """
    for r, cells in enumerate(rows):
        vals = [clean(c).lower() for c in cells]
        cols = {"flat": None, "room": None, "make": None, "cap": None}
        for i, v in enumerate(vals):
            if not v:
                continue
            if cols["room"] is None and v in H_ROOM:
                cols["room"] = i
            elif cols["make"] is None and v in H_MAKE:
                cols["make"] = i
            elif cols["cap"] is None and v in H_CAP:
                cols["cap"] = i
            elif cols["flat"] is None and v in H_FLAT:
                cols["flat"] = i
        if all(cols[k] is not None for k in ("room", "make", "cap")):
            return r, cols
    return None, None


def parse_residential(rows, prefix):
    hdr, cols = find_header(rows)
    if hdr is None:
        # No header row — the 165 shape: title row, then Flat|Room|Make|Cap
        hdr, cols = 0, {"flat": 0, "room": 1, "make": 2, "cap": 3}

    out = []
    for cells in rows[hdr + 1:]:
        def get(i):
            return clean(cells[i]) if i is not None and i < len(cells) else ""

        flat, room = get(cols["flat"]), get(cols["room"])
        make, cap = get(cols["make"]), get(cols["cap"])
        if not room or not make:
            continue
        if room.lower().startswith("total") or flat.lower().startswith("total"):
            continue

        room = fix_room(room)
        hp = hp_value(cap)
        if hp > PACKAGE_ABOVE_HP:
            category = "Package Unit"
        elif hp >= STANDING_FROM_HP:
            category = "Standing Unit"
        else:
            category = "Split AC"
        if flat:
            flat_t = fix_flat(flat)
            same = flat.strip().lower() == room.strip().lower()
            where = f"{prefix} · {flat_t}" if same else f"{prefix} · {flat_t} · {room}"
        else:
            where = f"{prefix} · {room}"
        out.append({"name": f"{where} — {make} {hp_text(cap)}", "category": category})
    return out


def parse_office(rows, prefix):
    out = []
    for cells in rows:
        c = [clean(x) for x in (list(cells) + [""] * 8)[:8]]
        sn, block, zone, area, brand, utype, hp, qty = c
        if not brand or not area:
            continue
        if sn.lower() == "s/n" or zone.lower().startswith("floor"):
            continue
        try:
            n = int(float(qty)) if qty else 1
        except ValueError:
            n = 1
        base = f"{prefix} · {short_zone(zone)} · {fix_room(area)} — {brand} {utype} {hp_text(hp)}"
        category = "Standing Unit" if "standing" in utype.lower() else "Split AC"
        for _ in range(max(n, 1)):
            out.append({"name": base, "category": category})
    return out


def number_duplicates(assets):
    """Repeated names mean separate machines. Number them in sheet order."""
    counts = {}
    for a in assets:
        counts[a["name"]] = counts.get(a["name"], 0) + 1
    seen = {}
    numbered = 0
    for a in assets:
        if counts[a["name"]] > 1:
            base = a["name"]
            seen[base] = seen.get(base, 0) + 1
            a["name"] = f"{base} #{seen[base]}"
            numbered += 1
    return assets, numbered


def read_sheet(path, sheet):
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"Sheet '{sheet}' not found in {path}. Available: {wb.sheetnames}")
    rows = []
    for row in wb[sheet].iter_rows(values_only=True):
        vals = list(row)
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        rows.append(vals)
    return rows


def main():
    plan = []
    for src in SOURCES:
        if not os.path.exists(src["file"]):
            print(f"!! missing file, skipped: {src['file']}")
            continue
        rows = read_sheet(src["file"], src["sheet"])
        parse = parse_residential if src["layout"] == "residential" else parse_office
        assets, numbered = number_duplicates(parse(rows, src["prefix"]))
        plan.append((src, assets))
        note = f"   {numbered} numbered" if numbered else ""
        print(f"{src['location']:<15} {len(rows):>3} rows  ->  {len(assets):>3} units{note}")

    print(f"\nTOTAL UNITS: {sum(len(a) for _, a in plan)}")

    print("\nSample of generated names:")
    for src, assets in plan:
        print(f"\n  [{src['location']}]")
        for a in assets[:3]:
            print(f"    {a['category']:<14} {a['name']}")
        if len(assets) > 3:
            print(f"    ... and {len(assets) - 3} more")

    if DRY_RUN:
        print("\nDRY_RUN is True — nothing written. Set DRY_RUN = False to import.")
        return

    url = os.environ.get("DATABASE_URL")
    if not url:
        sys.exit("DATABASE_URL not set.")

    conn = psycopg2.connect(url)
    conn.autocommit = False
    cur = conn.cursor()
    now = datetime.now(timezone.utc).replace(tzinfo=None)

    try:
        cur.execute('SELECT id, name FROM "Organization" ORDER BY "createdAt" LIMIT 2;')
        orgs = cur.fetchall()
        if len(orgs) != 1:
            sys.exit(f"Expected exactly one organisation, found {len(orgs)}: {orgs}")
        org_id, org_name = orgs[0]
        print(f"\nOrganisation: {org_name}")

        created = skipped = 0
        for src, assets in plan:
            cur.execute('SELECT id FROM "Location" WHERE lower(name) = lower(%s) LIMIT 1;',
                        (src["location"],))
            row = cur.fetchone()
            if row:
                loc_id = row[0]
                if FORCE_LOCATION_UPDATE:
                    cur.execute("""
                        UPDATE "Location" SET
                          client = COALESCE(%s, client),
                          "supervisorName"  = COALESCE(%s, "supervisorName"),
                          "supervisorPhone" = COALESCE(%s, "supervisorPhone"),
                          area = COALESCE(%s, area)
                        WHERE id = %s;
                    """, (src["client"], src["supervisor"], src["phone"], src["area"], loc_id))
                else:
                    cur.execute("""
                        UPDATE "Location" SET
                          client = COALESCE(client, %s),
                          "supervisorName"  = COALESCE("supervisorName", %s),
                          "supervisorPhone" = COALESCE("supervisorPhone", %s),
                          area = COALESCE(area, %s)
                        WHERE id = %s;
                    """, (src["client"], src["supervisor"], src["phone"], src["area"], loc_id))
                print(f"\n{src['location']}: existing location")
            else:
                loc_id = str(uuid.uuid4())
                cur.execute("""
                    INSERT INTO "Location"
                      (id, name, "organizationId", "createdAt",
                       client, "supervisorName", "supervisorPhone", area, "isActive")
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, true);
                """, (loc_id, src["location"], org_id, now,
                      src["client"], src["supervisor"], src["phone"], src["area"]))
                print(f"\n{src['location']}: location CREATED")

            new_here = 0
            for a in assets:
                cur.execute('SELECT 1 FROM "Asset" WHERE name = %s AND "locationId" = %s LIMIT 1;',
                            (a["name"], loc_id))
                if cur.fetchone():
                    skipped += 1
                    continue
                cur.execute("""
                    INSERT INTO "Asset"
                      (id, name, category, status, "organizationId", "locationId",
                       "createdAt", "updatedAt")
                    VALUES (%s, %s, %s, 'OPERATIONAL', %s, %s, %s, %s);
                """, (str(uuid.uuid4()), a["name"], a["category"], org_id, loc_id, now, now))
                created += 1
                new_here += 1
            print(f"  {new_here} new, {len(assets) - new_here} already present")

        cur.execute('SELECT count(*) FROM "Asset";')
        print(f"\nCreated: {created}   Already present: {skipped}")
        print(f"Assets in database after import: {cur.fetchone()[0]}")
        conn.commit()
        print("Committed.")
    except Exception as e:
        conn.rollback()
        print("Rolled back. Error:", e)
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
